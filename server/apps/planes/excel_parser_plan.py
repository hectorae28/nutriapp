"""
Parser de Excel de Plan Alimenticio.
Lee un archivo Excel con el formato del plan y retorna un diccionario estructurado.
"""
from datetime import datetime
from decimal import Decimal
import openpyxl


# Mapeo de número de lista a grupo
LISTA_A_GRUPO = {
    1: "LECHE",
    2: "VEGETALES",
    3: "FRUTAS",
    4: "ALMIDONES",
    5: "CARNES_A",
    6: "GRASAS",
}


def parsear_excel_plan(file_path_or_bytes):
    """
    Parsea un archivo Excel de plan alimenticio.
    
    Args:
        file_path_or_bytes: ruta al archivo Excel o bytes del archivo
    
    Returns:
        dict con estructura:
        {
            'plan': {
                'fecha_inicio': date,
                'kcal_objetivo': int,
                'pct_proteinas': Decimal,
                'pct_grasas': Decimal,
                'pct_carbohidratos': Decimal,
                'requerimiento_hidrico_ml': int,
                'notas': str,
                'tipo_dieta': str,
            },
            'tiempos_comida': [
                {
                    'nombre': str,
                    'hora': str,
                    'orden': int,
                    'raciones': [
                        {'grupo_nombre': str, 'cantidad': Decimal}
                    ]
                }
            ],
            'advertencias': []
        }
    """
    wb = openpyxl.load_workbook(file_path_or_bytes, data_only=True)
    advertencias = []
    
    # Buscar hojas
    ws_presentacion = None
    ws_listas = None
    
    for sheet_name in wb.sheetnames:
        if "presentacion" in sheet_name.lower() or "recomendaciones" in sheet_name.lower():
            ws_presentacion = wb[sheet_name]
        elif "lista" in sheet_name.lower() or "intercambio" in sheet_name.lower():
            ws_listas = wb[sheet_name]
    
    if not ws_presentacion and not ws_listas:
        # Intentar con las primeras 2 hojas
        if len(wb.sheetnames) >= 2:
            ws_presentacion = wb[wb.sheetnames[0]]
            ws_listas = wb[wb.sheetnames[1]]
        else:
            advertencias.append("No se encontraron las hojas esperadas")
            return {'plan': {}, 'tiempos_comida': [], 'advertencias': advertencias}
    
    plan_data = {}
    tiempos_comida = []
    
    # ── Parsear Sheet "Presentación" ──────────────────────────────────────────
    if ws_presentacion:
        # R8 col44: fecha (AQ8 en Excel = col 44)
        fecha_cell = ws_presentacion.cell(row=8, column=44).value
        if fecha_cell:
            if isinstance(fecha_cell, datetime):
                plan_data['fecha_inicio'] = fecha_cell.date()
            elif isinstance(fecha_cell, str):
                try:
                    plan_data['fecha_inicio'] = datetime.strptime(fecha_cell, "%d/%m/%Y").date()
                except:
                    try:
                        plan_data['fecha_inicio'] = datetime.strptime(fecha_cell, "%Y-%m-%d").date()
                    except:
                        advertencias.append(f"No se pudo parsear la fecha: {fecha_cell}")
        
        # R12 col46: kcal (parsear "1750/ día" → 1750)
        kcal_cell = ws_presentacion.cell(row=12, column=46).value
        if kcal_cell:
            if isinstance(kcal_cell, (int, float)):
                plan_data['kcal_objetivo'] = int(kcal_cell)
            elif isinstance(kcal_cell, str):
                # Extraer números
                import re
                numeros = re.findall(r'\d+', kcal_cell)
                if numeros:
                    plan_data['kcal_objetivo'] = int(numeros[0])
        
        # R12 col54: requerimiento hídrico
        hidrico_cell = ws_presentacion.cell(row=12, column=54).value
        if hidrico_cell:
            if isinstance(hidrico_cell, (int, float)):
                plan_data['requerimiento_hidrico_ml'] = int(hidrico_cell)
            elif isinstance(hidrico_cell, str):
                import re
                numeros = re.findall(r'\d+', hidrico_cell)
                if numeros:
                    plan_data['requerimiento_hidrico_ml'] = int(numeros[0])
        
        # R16: proteínas g/kg, grasas g/kg, carbohidratos g/kg
        # Estos son valores calculados. Para recuperar los porcentajes,
        # necesitaríamos hacer la operación inversa o buscar en otra celda.
        # Por simplicidad, usamos valores por defecto o los dejamos vacíos.
        # En un Excel real, podrías tener celdas específicas para los %.
        
        # R28+: notas
        notas_cell = ws_presentacion.cell(row=29, column=1).value
        if notas_cell:
            plan_data['notas'] = str(notas_cell)
    
    # ── Parsear Sheet "Listas de Intercambio" ─────────────────────────────────
    if ws_listas:
        # R5: COMIDA 1-7 (cols 29, 40, 51, 62, 73, 84, 95)
        # R6: nombres y horas
        # R7: headers "Raciones", "Lista"
        # R8-R13: datos de raciones
        
        # Detectar tiempos de comida
        col_inicio_tiempos = [29 + (i * 11) for i in range(7)]  # COMIDA 1-7
        
        for idx, col in enumerate(col_inicio_tiempos):
            # R6: nombre
            nombre_cell = ws_listas.cell(row=6, column=col).value
            if not nombre_cell:
                continue
            
            # R6 col+5: hora
            hora_cell = ws_listas.cell(row=6, column=col+5).value
            hora_str = ""
            if hora_cell:
                if isinstance(hora_cell, datetime):
                    hora_str = hora_cell.strftime("%H:%M")
                elif isinstance(hora_cell, str):
                    hora_str = hora_cell
            
            tiempo = {
                'nombre': str(nombre_cell).strip(),
                'hora': hora_str,
                'orden': idx,
                'raciones': []
            }
            
            # R8-R13: leer raciones (o más filas si hay más grupos)
            for fila in range(8, 20):  # Amplio rango para cubrir todos los grupos
                # Col de raciones
                cantidad_cell = ws_listas.cell(row=fila, column=col).value
                lista_cell = ws_listas.cell(row=fila, column=col+5).value
                
                if cantidad_cell and lista_cell:
                    try:
                        cantidad = Decimal(str(cantidad_cell))
                        lista_num = int(lista_cell)
                        grupo_nombre = LISTA_A_GRUPO.get(lista_num)
                        
                        if grupo_nombre:
                            tiempo['raciones'].append({
                                'grupo_nombre': grupo_nombre,
                                'cantidad': cantidad
                            })
                    except (ValueError, TypeError):
                        continue
            
            if tiempo['raciones']:
                tiempos_comida.append(tiempo)
    
    # Valores por defecto para porcentajes (si no se pueden extraer del Excel)
    if 'pct_proteinas' not in plan_data:
        plan_data['pct_proteinas'] = Decimal('20.0')
    if 'pct_grasas' not in plan_data:
        plan_data['pct_grasas'] = Decimal('30.0')
    if 'pct_carbohidratos' not in plan_data:
        plan_data['pct_carbohidratos'] = Decimal('50.0')
    
    return {
        'plan': plan_data,
        'tiempos_comida': tiempos_comida,
        'advertencias': advertencias
    }
