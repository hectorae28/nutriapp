"""
Generador Excel de Plan Alimenticio.
Crea un archivo Excel de 3 hojas siguiendo el formato de referencia:
- Sheet 1: Presentación - Recomendaciones
- Sheet 2: Listas de Intercambio
- Sheet 3: Ejemplo (menú tipo)
"""
import os
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

MEDICO_NOMBRE = os.environ.get("MEDICO_NOMBRE", "Dr. Domingo Porras")

# Mapeo de grupos a listas de intercambio
GRUPO_A_LISTA = {
    "LECHE": 1,
    "VEGETALES": 2,
    "FRUTAS": 3,
    "ALMIDONES": 4,
    "AZUCAR": 4,
    "CARNES_A": 5,
    "CARNES_B": 5,
    "GRASAS": 6,
}


def generar_excel_plan(plan) -> bytes:
    """
    Genera un archivo Excel con el plan alimenticio.
    
    Args:
        plan: PlanAlimenticio con select_related('paciente__user') y
              prefetch_related('tiempos_comida__raciones__grupo', 'alimento_tags__alimento')
    
    Returns:
        bytes: contenido del archivo Excel
    """
    wb = Workbook()
    
    # Eliminar la hoja por defecto
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    # Crear las 3 hojas
    _crear_hoja_presentacion(wb, plan)
    _crear_hoja_listas(wb, plan)
    _crear_hoja_ejemplo(wb, plan)
    
    # Guardar en BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _crear_hoja_presentacion(wb, plan):
    """Crea la hoja 'Presentación - Recomendaciones'"""
    ws = wb.create_sheet("Presentación - Recomendaciones", 0)
    
    # Obtener datos del plan y paciente
    paciente = plan.paciente
    expediente = getattr(paciente, 'expediente', None)
    peso_actual = expediente.peso_usual_kg if expediente else None
    
    # Calcular próxima visita (ejemplo: +30 días)
    fecha_inicio = plan.fecha_inicio or date.today()
    proxima_visita = fecha_inicio + timedelta(days=30)
    
    # R7: FECHA y PRÓXIMA VISITA (cols 44=AR y 57=BE)
    ws.cell(row=7, column=44, value="FECHA:")
    ws.cell(row=7, column=44).font = Font(bold=True)
    ws.cell(row=7, column=57, value="PRÓXIMA VISITA:")
    ws.cell(row=7, column=57).font = Font(bold=True)
    
    # R8: fechas
    ws.cell(row=8, column=44, value=fecha_inicio.strftime("%d/%m/%Y"))
    ws.cell(row=8, column=57, value=proxima_visita.strftime("%d/%m/%Y"))
    
    # R11: labels
    ws.cell(row=11, column=1, value="PACIENTE:")
    ws.cell(row=11, column=1).font = Font(bold=True)
    ws.cell(row=11, column=37, value="PESO ACTUAL:")
    ws.cell(row=11, column=37).font = Font(bold=True)
    ws.cell(row=11, column=46, value="KCAL REQUERIDAS:")
    ws.cell(row=11, column=46).font = Font(bold=True)
    ws.cell(row=11, column=54, value="REQUERIMIENTO HÍDRICO:")
    ws.cell(row=11, column=54).font = Font(bold=True)
    
    # R12: datos
    nombre_paciente = paciente.user.get_full_name() or paciente.user.username
    ws.cell(row=12, column=1, value=nombre_paciente)
    ws.cell(row=12, column=37, value=f"{peso_actual} kg" if peso_actual else "—")
    ws.cell(row=12, column=46, value=f"{plan.kcal_objetivo}/ día" if plan.kcal_objetivo else "—")
    ws.cell(row=12, column=54, value=f"{plan.requerimiento_hidrico_ml} ml" if plan.requerimiento_hidrico_ml else "—")
    
    # R15: labels de macros
    ws.cell(row=15, column=1, value="PROTEÍNAS:")
    ws.cell(row=15, column=1).font = Font(bold=True)
    ws.cell(row=15, column=12, value="GRASAS:")
    ws.cell(row=15, column=12).font = Font(bold=True)
    ws.cell(row=15, column=20, value="CARBOHIDRATOS:")
    ws.cell(row=15, column=20).font = Font(bold=True)
    ws.cell(row=15, column=28, value="COMPLEMENTO:")
    ws.cell(row=15, column=28).font = Font(bold=True)
    
    # R16: valores calculados de macros
    vct = float(plan.kcal_objetivo or 0)
    peso = float(peso_actual or 0)
    
    if vct > 0 and peso > 0:
        pct_p = float(plan.pct_proteinas or 0)
        pct_g = float(plan.pct_grasas or 0)
        pct_cho = float(plan.pct_carbohidratos or 0)
        
        g_p = (vct * pct_p / 100) / 4
        g_kg_p = g_p / peso
        
        g_g = (vct * pct_g / 100) / 9
        g_kg_g = g_g / peso
        
        g_cho = (vct * pct_cho / 100) / 4
        g_kg_cho = g_cho / peso
        
        ws.cell(row=16, column=1, value=f"{g_kg_p:.1f} g/Kg-p/día")
        ws.cell(row=16, column=12, value=f"{g_kg_g:.1f} g/Kg-p/día")
        ws.cell(row=16, column=20, value=f"{g_kg_cho:.1f} g/Kg-p/día")
    
    ws.cell(row=16, column=28, value=plan.suplemento_complemento or "—")
    
    # R28: RECOMENDACIONES Y OBSERVACIONES
    ws.cell(row=28, column=1, value="RECOMENDACIONES Y OBSERVACIONES:")
    ws.cell(row=28, column=1).font = Font(bold=True)
    ws.cell(row=29, column=1, value=plan.notas or "")
    
    # Footer (última fila)
    ws.cell(row=35, column=1, value=MEDICO_NOMBRE)
    ws.cell(row=35, column=1).font = Font(italic=True, size=9)


def _crear_hoja_listas(wb, plan):
    """Crea la hoja 'Listas de Intercambio'"""
    ws = wb.create_sheet("Listas de Intercambio", 1)
    
    # R3: título
    ws.cell(row=3, column=29, value="PLAN DE ALIMENTACIÓN SUGERIDO")
    ws.cell(row=3, column=29).font = Font(bold=True, size=12)
    ws.cell(row=3, column=29).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=3, start_column=29, end_row=3, end_column=50)
    
    # Obtener tiempos de comida ordenados
    tiempos = list(plan.tiempos_comida.order_by("orden"))
    
    # R5: COMIDA 1, COMIDA 2... (cada una ocupa 11 cols empezando en col 29)
    for idx, tc in enumerate(tiempos):
        col_inicio = 29 + (idx * 11)
        ws.cell(row=5, column=col_inicio, value=f"COMIDA {idx+1}")
        ws.cell(row=5, column=col_inicio).font = Font(bold=True)
        ws.merge_cells(start_row=5, start_column=col_inicio, end_row=5, end_column=col_inicio+10)
        ws.cell(row=5, column=col_inicio).alignment = Alignment(horizontal="center")
    
    # R6: nombre del tiempo + hora
    for idx, tc in enumerate(tiempos):
        col_inicio = 29 + (idx * 11)
        hora_str = tc.hora.strftime("%H:%M") if tc.hora else ""
        ws.cell(row=6, column=col_inicio, value=tc.nombre)
        ws.cell(row=6, column=col_inicio+5, value=hora_str)
    
    # R7: headers "Raciones", "Lista"
    for idx in range(len(tiempos)):
        col_inicio = 29 + (idx * 11)
        ws.cell(row=7, column=col_inicio, value="Raciones")
        ws.cell(row=7, column=col_inicio).font = Font(bold=True, size=9)
        ws.cell(row=7, column=col_inicio+5, value="Lista")
        ws.cell(row=7, column=col_inicio+5).font = Font(bold=True, size=9)
    
    # R8+: datos de raciones por grupo
    # Obtener todos los grupos únicos
    grupos_dict = {}
    for tc in tiempos:
        for racion in tc.raciones.all():
            nombre_grupo = racion.grupo.nombre
            if nombre_grupo not in grupos_dict:
                grupos_dict[nombre_grupo] = racion.grupo
    
    # Orden de grupos (puedes personalizar)
    grupos_ordenados = ["LECHE", "VEGETALES", "FRUTAS", "ALMIDONES", "CARNES_A", "CARNES_B", "GRASAS", "AZUCAR"]
    grupos_presentes = [g for g in grupos_ordenados if g in grupos_dict]
    
    # Escribir datos
    fila_actual = 8
    for grupo_nombre in grupos_presentes:
        # Nombre del grupo en col 1
        ws.cell(row=fila_actual, column=1, value=grupo_nombre)
        ws.cell(row=fila_actual, column=1).font = Font(bold=True, size=9)
        
        # Datos por cada tiempo
        for idx, tc in enumerate(tiempos):
            col_inicio = 29 + (idx * 11)
            # Buscar racion para este grupo en este tiempo
            racion = next((r for r in tc.raciones.all() if r.grupo.nombre == grupo_nombre), None)
            if racion:
                cantidad = float(racion.cantidad)
                ws.cell(row=fila_actual, column=col_inicio, value=cantidad)
                # Lista
                num_lista = GRUPO_A_LISTA.get(grupo_nombre, 0)
                ws.cell(row=fila_actual, column=col_inicio+5, value=num_lista)
        
        fila_actual += 1
    
    # R15-R90: LISTAS DE INTERCAMBIO (tabla estática de referencia)
    # Esta es una tabla fija que siempre es igual. Por simplicidad, agregamos un placeholder.
    ws.cell(row=15, column=1, value="LISTAS DE INTERCAMBIO (Tabla de referencia)")
    ws.cell(row=15, column=1).font = Font(bold=True, size=11)
    ws.cell(row=16, column=1, value="[Aquí iría la tabla estática de listas de intercambio con todos los alimentos]")
    ws.cell(row=16, column=1).font = Font(italic=True, size=9)


def _crear_hoja_ejemplo(wb, plan):
    """Crea la hoja 'Ejemplo' con menú tipo"""
    ws = wb.create_sheet("Ejemplo", 2)
    
    # R6: título
    ws.cell(row=6, column=1, value="MENÚ TIPO (EJEMPLO DE COMBINACIÓN DE ALIMENTOS)")
    ws.cell(row=6, column=1).font = Font(bold=True, size=12)
    
    # Verificar si hay alimento_tags
    alimento_tags = list(plan.alimento_tags.select_related('alimento__grupo').all())
    
    if not alimento_tags:
        ws.cell(row=8, column=1, value="No se han definido alimentos específicos para este plan.")
        ws.cell(row=8, column=1).font = Font(italic=True)
        return
    
    # Agrupar alimentos por tag
    fila_actual = 8
    
    # Headers
    ws.cell(row=fila_actual, column=1, value="ETIQUETA")
    ws.cell(row=fila_actual, column=2, value="ALIMENTO")
    ws.cell(row=fila_actual, column=3, value="GRUPO")
    ws.cell(row=fila_actual, column=4, value="NOTA")
    for col in range(1, 5):
        ws.cell(row=fila_actual, column=col).font = Font(bold=True, size=9)
    fila_actual += 1
    
    # Datos de alimentos etiquetados
    for at in alimento_tags:
        ws.cell(row=fila_actual, column=1, value=at.get_tag_display())
        ws.cell(row=fila_actual, column=2, value=at.alimento.nombre if at.alimento else "")
        ws.cell(row=fila_actual, column=3, value=at.alimento.grupo.nombre if at.alimento and at.alimento.grupo else "")
        ws.cell(row=fila_actual, column=4, value=at.nota or "")
        fila_actual += 1
