import io
import os
from datetime import date
from django.db.models import Q
from django.http import HttpResponse
from rest_framework import permissions, viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from django.contrib.auth import get_user_model
from django.db import transaction

from .models import Paciente
from .permissions import IsNutricionista, IsNutricionistaOrPaciente
from .serializers import PacienteSerializer
from .excel_parser import parse_historia_nutricional
from apps.expediente.models import ExpedienteClinico, ConsumoCaloricoItem, ExamenBioquimico, RegistroProgreso
from apps.expediente.serializers import ExpedienteClinicoSerializer, ExamenBioquimicoSerializer

# ── Mapeo de claves del parser → campos del modelo ExamenBioquimico ──────────
# CORRECCIÓN: Ahora el parser devuelve UN SOLO diccionario con TODOS los campos
# (ambas tablas visuales combinadas). Usar un solo mapa completo.
_EXAM_FIELD_MAP = {
    # Campos de Tabla 1 (fila 174)
    'pt': 'proteinas_totales', 'alb': 'albumina', 'glo': 'globulina',
    'ur': 'urea', 'cr': 'creatinina', 'au': 'acido_urico',
    'col': 'colesterol', 'hdl': 'hdl', 'ldl': 'ldl', 'vldl': 'vldl',
    'tgc': 'trigliceridos', 'gli': 'glucosa', 'gli_p': 'glucosa_postprandial',
    'insulina': 'insulina', 'insulina_p': 'insulina_postprandial',
    # Campos de Tabla 2 (fila 190)
    'hba': 'hemoglobina_glicosilada', 'tgo': 'tgo', 'tgp': 'tgp',
    'bl_t': 'bilirrubina_total', 'bl_d': 'bilirrubina_directa', 'bl_i': 'bilirrubina_indirecta',
    'hb': 'hemoglobina', 'hct': 'hematocrito',
    't3': 't3', 't4': 't4', 'tsh': 'tsh',
    'fe': 'hierro', 'b12': 'vitamina_b12',
    'na': 'sodio', 'potasio': 'potasio', 'cloro': 'cloro', 'calcio': 'calcio',
}


def _traducir_examen(raw_dict):
    """Traduce un dict crudo del parser a campos del modelo ExamenBioquimico.
    Ahora usa un solo mapa que incluye TODOS los campos (ambas tablas combinadas)."""
    translated = {'fecha': raw_dict.get('fecha')}
    for parser_key, model_field in _EXAM_FIELD_MAP.items():
        val = raw_dict.get(parser_key)
        if val is not None:
            translated[model_field] = val
    return translated


class PacienteViewSet(viewsets.ModelViewSet):
    serializer_class = PacienteSerializer
    permission_classes = [permissions.IsAuthenticated, IsNutricionistaOrPaciente]

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser or user.groups.filter(name__in=["Nutricionista"]).exists():
            qs = Paciente.objects.select_related("user").all()
        else:
            # Paciente solo ve el suyo
            qs = Paciente.objects.select_related("user").filter(user=user)

        search = self.request.query_params.get("search", "")
        if search:
            qs = qs.filter(
                Q(user__first_name__icontains=search)
                | Q(user__last_name__icontains=search)
                | Q(cedula__icontains=search)
                | Q(user__username__icontains=search)
            )
        return qs

    def get_permissions(self):
        if self.action in ["create", "destroy"]:
            return [permissions.IsAuthenticated(), IsNutricionista()]
        return super().get_permissions()


class PacienteExcelImportView(APIView):
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [IsNutricionista]

    def post(self, request, *args, **kwargs):
        if 'archivo' not in request.FILES:
            return Response(
                {"error": "No se ha proporcionado ningún archivo."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['archivo']

        # Conservar la extensión original (.xls o .xlsx) para que el parser la detecte
        _, ext = os.path.splitext(file.name)
        ext = ext.lower() if ext else '.xls'
        tmp_path = f'/tmp/uploaded_excel{ext}'

        with open(tmp_path, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)

        try:
            parsed_data = parse_historia_nutricional(tmp_path)
        except Exception as e:
            return Response(
                {"error": f"Error al parsear el archivo Excel: {e}", "advertencias": []},
                status=status.HTTP_400_BAD_REQUEST
            )
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

        if not parsed_data.get('paciente', {}).get('cedula'):
            return Response(
                {"error": "No se pudo extraer la cédula del paciente del archivo.", "advertencias": parsed_data['advertencias']},
                status=status.HTTP_400_BAD_REQUEST
            )

        User = get_user_model()
        paciente_data = parsed_data['paciente']
        expediente_data = parsed_data['expediente']
        consumo_calorico_data = parsed_data['consumo_calorico']
        examenes_data = parsed_data['examenes']
        registros_progreso_data = parsed_data.get('registros_progreso', [])
        tratamiento_data = parsed_data.get('tratamiento', {})
        plan_grupos_data = parsed_data.get('plan_grupos', [])
        recordatorio_data = parsed_data.get('recordatorio', {})

        paciente_instance = None
        created_user = False
        
        try:
            with transaction.atomic():
                # 1. Get or create User and Paciente
                cedula = paciente_data['cedula']
                user_email = paciente_data.get('email', f'paciente_{cedula}@nutriapp.com')
                user_username = f'paciente_{cedula}'

                # Try to find Paciente by cedula
                try:
                    paciente_instance = Paciente.objects.get(cedula=cedula)
                    user_instance = paciente_instance.user
                    # Update user fields if needed
                    user_instance.first_name = paciente_data.get('nombre', user_instance.first_name)
                    user_instance.last_name = paciente_data.get('apellido', user_instance.last_name)
                    user_instance.email = user_email
                    user_instance.save()

                except Paciente.DoesNotExist:
                    # If Paciente does not exist, check if User with same email/username exists
                    try:
                        user_instance = User.objects.get(Q(email=user_email) | Q(username=user_username))
                    except User.DoesNotExist:
                        # Create User if not found
                        user_instance = User.objects.create_user(
                            username=user_username,
                            email=user_email,
                            password=None, # Sin password, se establece como inutilizable
                            first_name=paciente_data.get('nombre', ''),
                            last_name=paciente_data.get('apellido', '')
                        )
                        created_user = True
                        # Add user to 'Paciente' group
                        from django.contrib.auth.models import Group
                        paciente_group, _ = Group.objects.get_or_create(name='Paciente')
                        user_instance.groups.add(paciente_group)
                    
                    # Create Paciente
                    paciente_instance = Paciente.objects.create(user=user_instance, **{
                        k: v for k, v in paciente_data.items() if k not in ['nombre', 'apellido', 'email']
                    })

                # Update Paciente fields
                for attr, value in paciente_data.items():
                    if attr not in ['nombre', 'apellido', 'email']:
                        setattr(paciente_instance, attr, value)
                paciente_instance.save()

                # 2. Get or create ExpedienteClinico
                expediente_instance, created_expediente = ExpedienteClinico.objects.get_or_create(
                    paciente=paciente_instance
                )

                # Update ExpedienteClinico using its serializer
                expediente_serializer = ExpedienteClinicoSerializer(
                    expediente_instance, data=expediente_data, partial=True
                )
                expediente_serializer.is_valid(raise_exception=True)
                expediente_serializer.save()

                # 3. Handle ConsumoCaloricoItem (delete old, create new)
                expediente_instance.consumo_calorico.all().delete() # Delete existing items
                for item_data in consumo_calorico_data:
                    ConsumoCaloricoItem.objects.create(expediente=expediente_instance, **item_data)

                # 4. Handle ExamenBioquimico — traducir claves del parser → campos del modelo
                # CORRECCIÓN: Ahora examenes_data contiene UN SOLO registro (ambas tablas combinadas)
                for raw_exam in examenes_data:
                    translated = _traducir_examen(raw_exam)
                    if not translated.get('fecha'):
                        translated['fecha'] = date.today()
                    translated['paciente'] = paciente_instance
                    # Solo crear si hay al menos un valor lab (no solo fecha/paciente)
                    has_values = any(
                        v is not None for k, v in translated.items()
                        if k not in ('fecha', 'paciente')
                    )
                    if has_values:
                        ExamenBioquimico.objects.create(**translated)

                # 5. Handle RegistroProgreso (Antropometría) — crear registros nuevos
                for reg_data in registros_progreso_data:
                    RegistroProgreso.objects.get_or_create(
                        paciente=paciente_instance,
                        fecha=reg_data['fecha'],
                        defaults={
                            'peso_kg': reg_data['peso_kg'],
                            'talla_cm': reg_data.get('talla_cm'),
                            'creado_por': request.user,
                        }
                    )

                # 6. Handle RecordatorioAlimentario — crear con entradas
                if recordatorio_data.get('entradas'):
                    from apps.expediente.models import RecordatorioAlimentario, EntradaRecordatorio
                    # Eliminar recordatorios previos del paciente (si se reimporta)
                    RecordatorioAlimentario.objects.filter(paciente=paciente_instance).delete()
                    
                    # Crear recordatorio con fecha de hoy
                    recordatorio = RecordatorioAlimentario.objects.create(
                        paciente=paciente_instance,
                        fecha=date.today(),
                        creado_por=request.user
                    )
                    
                    # Crear entradas
                    for entrada_data in recordatorio_data['entradas']:
                        EntradaRecordatorio.objects.create(
                            recordatorio=recordatorio,
                            nombre=entrada_data.get('nombre', ''),
                            hora=entrada_data.get('hora'),
                            descripcion=entrada_data.get('descripcion', ''),
                            orden=entrada_data.get('orden', 0)
                        )

                # 7. NO crear PlanAlimenticio durante importación Excel
                # Solo se deben crear/actualizar: User + Paciente + ExpedienteClinico
                # + ConsumoCaloricoItem[] + ExamenBioquimico[] + RegistroProgreso[]
                # + RecordatorioAlimentario + EntradaRecordatorio[]

                return Response(
                    {
                        "message": "Importación de datos completada exitosamente.",
                        "paciente_id": paciente_instance.id,
                        "paciente_nombre": paciente_instance.user.get_full_name(),
                        "campos_extraidos": (
                            len(paciente_data) + len(expediente_data) +
                            len(consumo_calorico_data) * len(consumo_calorico_data[0] if consumo_calorico_data else []) +
                            len(examenes_data) * len(examenes_data[0] if examenes_data else [])
                        ),
                        "advertencias": parsed_data['advertencias'],
                    },
                    status=status.HTTP_200_OK
                )

        except Exception as e:
            return Response(
                {"error": f"Error al guardar los datos en la base de datos: {e}", "advertencias": parsed_data['advertencias']},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class PacienteExcelExportView(APIView):
    """GET /api/pacientes/{id}/exportar-excel/ — genera .xlsx con historia nutricional del paciente."""

    permission_classes = [IsNutricionista]

    def get(self, request, paciente_id, *args, **kwargs):
        try:
            paciente = Paciente.objects.select_related("user").get(pk=paciente_id)
        except Paciente.DoesNotExist:
            return Response({"error": "Paciente no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response({"error": "openpyxl no disponible"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historia Adultos"

        # ── Encabezado ────────────────────────────────────────────────────────
        header_font = Font(bold=True)
        ws.append(["HISTORIA NUTRICIONAL"])
        ws["A1"].font = Font(bold=True, size=14)

        ws.append([])

        # ── Datos Personales ──────────────────────────────────────────────────
        ws.append(["DATOS PERSONALES"])
        ws["A3"].font = header_font
        ws.append(["Paciente", paciente.user.get_full_name() or paciente.user.username])
        ws.append(["Cédula", paciente.cedula or ""])
        ws.append(["Fecha Nac.", str(paciente.fecha_nacimiento or "")])
        ws.append(["Sexo", paciente.get_sexo_display() if hasattr(paciente, 'get_sexo_display') else paciente.sexo or ""])
        ws.append(["Teléfono", paciente.telefono or ""])
        ws.append(["E-mail", paciente.user.email or ""])

        ws.append([])

        # ── Expediente Clínico ────────────────────────────────────────────────
        try:
            exp = ExpedienteClinico.objects.get(paciente=paciente)
            ws.append(["EXPEDIENTE CLÍNICO"])
            ws.cell(row=ws.max_row, column=1).font = header_font
            ws.append(["Motivo consulta", exp.motivo_consulta or ""])
            ws.append(["Antecedentes personales", exp.antecedentes_personales or ""])
            ws.append(["Antecedentes familiares", exp.antecedentes_familiares or ""])
            ws.append(["Peso máx (kg)", str(exp.peso_maximo_kg or "")])
            ws.append(["Peso mín (kg)", str(exp.peso_minimo_kg or "")])
            ws.append(["Peso usual (kg)", str(exp.peso_usual_kg or "")])
            ws.append(["Peso ideal (kg)", str(exp.peso_ideal_kg or "")])
            ws.append(["Peso deseado (kg)", str(exp.peso_deseado_kg or "")])
            ws.append(["Contextura", exp.contextura or ""])
            ws.append(["Observaciones calorías", exp.observaciones_calorias or ""])
            ws.append([])

            # Consumo calórico
            items_cc = exp.consumo_calorico.all()
            if items_cc.exists():
                ws.append(["CONSUMO CALÓRICO DIARIO"])
                ws.cell(row=ws.max_row, column=1).font = header_font
                ws.append(["Grupo", "INT", "P (g)", "G (g)", "CHO (g)", "KCAL"])
                for item in items_cc:
                    ws.append([item.get_grupo_display(), str(item.intercambios),
                                str(item.proteinas_g), str(item.grasas_g),
                                str(item.cho_g), str(item.kcal)])
                ws.append([])
        except ExpedienteClinico.DoesNotExist:
            pass

        # ── Exámenes Bioquímicos ──────────────────────────────────────────────
        examenes = ExamenBioquimico.objects.filter(paciente=paciente).order_by("-fecha")
        if examenes.exists():
            ws.append(["EXÁMENES BIOQUÍMICOS"])
            ws.cell(row=ws.max_row, column=1).font = header_font
            ws.append(["Fecha", "GLU", "COL", "HDL", "LDL", "TGC", "HBA1C", "TSH", "HB"])
            for ex in examenes:
                ws.append([
                    str(ex.fecha),
                    str(ex.glucosa or ""),
                    str(ex.colesterol or ""),
                    str(ex.hdl or ""),
                    str(ex.ldl or ""),
                    str(ex.trigliceridos or ""),
                    str(ex.hemoglobina_glicosilada or ""),
                    str(ex.tsh or ""),
                    str(ex.hemoglobina or ""),
                ])

        # ── Ajustar ancho de columnas ─────────────────────────────────────────
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        # ── Serializar y enviar ───────────────────────────────────────────────
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        nombre = (paciente.user.get_full_name() or paciente.user.username).replace(" ", "_")
        filename = f"historia_{nombre}_{paciente_id}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
